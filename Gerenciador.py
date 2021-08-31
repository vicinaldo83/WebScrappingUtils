from typing import List, Sequence, Union

import psutil as ps
from subprocess import Popen


class AppManager:
    def __init__(self, app:str) -> None:
        """
            parametros:
                app:str -> Caminho do executável 
                
            Abre o aplicativo passado em um processo assíncriono 
        """
        self.exec = Popen(app, shell=True)
    
    
    def Fechar(*args:Sequence[str]) -> None:
        """Finaliza o aplicativos passados"""
        
        for proc in ps.process_iter():
            if(isinstance(proc, ps.Process) and proc.name() in args):
                proc.kill()
        
        
    def Aguardar(self, app:str, fechar:bool=True, timeout:Union[None, int]=None) -> bool: 
        """
        Espera até o processo informado estar presente, se fechar for verdaderio [padrao], fecha o processo
        caso o contrário, apenas retorna verdadeiro. Caso não seja definido em limite de tempo [padrao], ele
        segue esperando indefinidamente
        """
        from time import sleep
        
        def do():
            for proc in self.Procurar(app):
                if(fechar): proc.kill()
                return True
            
            sleep(1)
            return False
        
        if(timeout is not None):
            for _ in range(timeout):
                if(do()): return True
            else: return False
        
        else:
            while(True):
                if(do()): return True


    def Finalizar(self, *args:Sequence[str]) -> None:
        """Finaliza o aplicativo iniciado junto ao processos passados"""
        
        self.exec.kill()
        self.Fechar(*args)
    
    
    def Procurar(*args:Sequence[str]) -> Union[List[ps.Process], List]:
        "Procura pelos processos informados e retorna os encontrados"
        
        process = []
        for proc in ps.process_iter():
            if(isinstance(proc, ps.Process) and proc.name() in args):
                process.append(proc)
        
        return process
                
    
    def __del__(self):
        try: self.finalizar()
        except: pass



#? Gerenciador dos arquivos de execução para evitar alterar muito a planilha
from os import makedirs
from Source.Config import PATH_PLAN, PATH_INFOS, BRR

class PlanManager:
    for arq in (PATH_PLAN, PATH_INFOS): makedirs(arq, exist_ok=True)
    
    def AtualizarJsonInfo(self, path:str=PATH_PLAN) -> Union[bool, dict]:
        "Le a planilha informada e cria as informaçoes para executar"
        
        from glob import glob
        from os.path import dirname, isdir, exists
        
        try:
            if(not isinstance(path, bytes)): 
                if(isdir(path)):
                    try: 
                        path = glob(dirname(path) + BRR + "*")[0]
                    except IndexError:
                        return (False, { 'error': f"Arquivo {path} não encontrado" })
                
                if(not exists(path)):
                    return (False, { 'error': f"Arquivo {path} não encontrado" })
                
            infos = self.LerPlanilha(path)
            if(infos):
                return (True, self.SalvarJsonInfo(infos))
            else: return (False, "Nenhuma informacao foi retornada")
        
        except Exception as e:
            return (False, { 'error': e })
        
    def LerJsonInfo(self, path:str=PATH_INFOS+"Jsons"+BRR+"Infos.json") -> dict:
        "Le as informacoes do arquivo json e retorna o dicionario"
        
        from json import load
        from os.path import isfile
        
        if(isfile(path)):
            with open(path, 'r', encoding='utf-8', errors='ignore') as file:
                info = load(file)
            
            return info if(info and isinstance(info, dict)) else {}
        else: return {}
    
    def SalvarJsonInfo(self, info:dict, path:str=PATH_INFOS+"Jsons"+BRR+"Infos.json") -> dict:
        "Recebe um dicionario e salva o json no caminho informado. Retorna o dicionario passado"
        
        from json import dump
        from os.path import dirname
        
        makedirs(dirname(path), exist_ok=True)
        with open(path, 'w', encoding='utf-8', errors='ignore') as file:
            dump(info, file, ensure_ascii=False, indent=2)
        
        return info
    

    def CopiarPlanilha(self, path:str) -> str:
        "Copia o planilha passada e salva em um lugar temporario"
        
        from os import remove
        from shutil import copyfile
        from os.path import dirname, exists
        
        newPath = PATH_INFOS + "Planilhas" + BRR + "infos.xlsx"
        makedirs(dirname(newPath), exist_ok=True)
        if(isinstance(path, bytes)):
            with open(newPath, 'wb') as file:
                file.write(path)
            
            return newPath
        
        makedirs(dirname(path), exist_ok=True)
        if(exists(newPath)):
            try:
                remove(newPath)
            except: pass

        copyfile(path, newPath)
        return newPath
        
    def LerPlanilha(self,
                    pathPlan:str,
                    pathJson:str=PATH_INFOS+"Jsons"+BRR+"Infos.json",
                    atualizar:bool=False) -> dict:
        """
            Abre e le todas as planilhas do diretorio passado no options.env, depois salva as infos.
            Caso atualizar seja verdadeiro [padrao e falso], o arquivo json passado ja criado e lido
            para ter suas informacoes atualizadas.
        """
        
        from datetime import datetime
        from openpyxl import load_workbook
        
        jsonEmpresas = self.LerJsonInfo(pathJson) if(atualizar) else {}
        try:
            wk = load_workbook(filename=self.CopiarPlanilha(pathPlan), read_only=True)
            tabela = wk.active
            linha = 1
            while(True):
                if(not tabela['A'+str(linha)].value): break
                elif('cod' in str(tabela['A'+str(linha)].value).lower()): linha += 1; continue
                else:
                    cod = str(tabela['A'+str(linha)].value)
                    jsonEmpresas.update({
                        cod + f'({linha})' if(jsonEmpresas.get(cod, False)) else cod:
                        {
                            'Nome' : str(tabela['B'+str(linha)].value),
                            'CPF-CNPJ': str(tabela['C'+str(linha)].value),
                            'Vencimento': tabela['D'+str(linha)].value.strftime("%d/%m/%Y") \
                                          if(isinstance(tabela['D'+str(linha)].value, datetime)) \
                                          else str(tabela['D'+str(linha)].value),
                            'Emissao' : None,
                            'Nota' : {
                                #True = Nota Fiscal, False = Nota de Despesas
                                'Tipo': bool(tabela['E'+str(linha)].value),
                                'Carteira': bool(tabela['G'+str(linha)].value),
                                'Numero': None,
                                
                                'ValorF': None,
                                'ValorExtraF': None,
                                
                                'ValorD': None,
                                'ValorExtraD': None,
                            },
                            'Status' : 1,
                            'Linha': linha
                        }
                    })
                    linha += 1;
        except Exception as e: print(e)
        finally:
            try:
                wk.close()
            except: pass
            
        return jsonEmpresas

    def SalvarPlanilha(self, infos:dict, path:str=PATH_INFOS + "Planilhas" + BRR + "infos.xlsx") -> bool:
        """
            Salva as informacoes do json dentro da planilha passada. Caso ela não exista, uma
            nova é criada
        """
        from os import startfile
        from datetime import datetime
        from os.path import exists, dirname
        
        makedirs(dirname(path), exist_ok=True)
        try:
            if(exists(path)):
                from openpyxl import load_workbook
                
                planilha = load_workbook(path)
                sheet = planilha.active
                
            else:
                from openpyxl import Workbook
                
                planilha = Workbook()
                sheet = planilha.active
                sheet.title = "Empresas"
            
            alfabeto = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
            from openpyxl.styles import PatternFill, Font, Side, Border
            
            thin = Side(border_style="thin", color="000000")
            for i in alfabeto:
                for c in sheet[i]:
                    c.fill = PatternFill(start_color="ffffff", end_color="ffffff")
                    c.font = Font("Arial", sz=9)
                    c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            
            grayTitle = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type='solid')
            for i, linha in enumerate(['COD', 'RAZAO SOCIAL', 'CPF-CNPJ', 'VENCIMENTO',
                                       'NF', 'ND', 'CARTEIRA', 'Nº NF OU ND',
                                       'VALOR NF', 'DIFERENCA',  'VALOR ND', 'DIFERENCA']):
                                
                sheet[alfabeto[i]+'1'].value = linha
                sheet[alfabeto[i]+'1'].fill = grayTitle
                sheet[alfabeto[i]+'1'].font = Font('Arial', b=True, sz=7)
            
            for key in infos.keys():
                try:
                    sheet['A' + str(infos[key]['Linha'])].value = int(key) #Cod
                except:
                    sheet['A' + str(infos[key]['Linha'])].value = key #Cod
                sheet['B' + str(infos[key]['Linha'])].value = infos[key]['Nome'] #Nome da empresa
                
                sheet['C'+ str(infos[key]['Linha'])].value = infos[key]['CPF-CNPJ'] #CPF ou CNPJ
                
                data = [int(x) for x in infos[key]['Vencimento'].split('/')]
                sheet['D' + str(infos[key]['Linha'])].value = datetime(day=data[0],
                                                                       month=data[1],
                                                                       year=data[2]) #Dia de vencimento
                
                #Nota Fiscal
                sheet['E' + str(infos[key]['Linha'])].value = "X" if(infos[key]['Nota']['Tipo']) else None
                #Nota de Despesas
                sheet['F' + str(infos[key]['Linha'])].value = None if(infos[key]['Nota']['Tipo']) else "X"
                #Se possui carteira
                sheet['G' + str(infos[key]['Linha'])].value = "X" if(infos[key]['Nota']['Carteira']) else None
                
                numero = infos[key]['Nota']['Numero']
                sheet['H' + str(infos[key]['Linha'])].value = str(int(numero)) if(numero) else numero #Numero da nota
                
                sheet['I' + str(infos[key]['Linha'])] = infos[key]['Nota']['ValorF'] #Valor da Nota Fiscal
                sheet['J' + str(infos[key]['Linha'])] = infos[key]['Nota']['ValorExtraF'] #Valor da diferença
                
                sheet['K' + str(infos[key]['Linha'])] = infos[key]['Nota']['ValorD'] #Valor da Nota de Despesas
                sheet['L' + str(infos[key]['Linha'])] = infos[key]['Nota']['ValorExtraD'] #Valor da diferença
                       
            from openpyxl.styles import Alignment           
            for i in alfabeto:
                for c in sheet[i]:
                    c.alignment = Alignment(shrinkToFit=True)
             
            planilha.save(path)
            startfile(path)
            return True
        except Exception as e:
            print(e)
            return False
        
